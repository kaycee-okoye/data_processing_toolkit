"""
    Module conatining a function for plotting data from a selected excel file.
    A list of PlotInfo objects corresponding to the desired sheetnames
    are used to provide formatting information for the figures.

    NOTE
    The input source for plot data is from an excel file. However, this can easily be
    modified for the users specific use case.
"""

from os.path import join
from openpyxl import load_workbook
from matplotlib import pyplot as plt
import numpy as np
import data_processing_toolkit.constants as constants
from data_processing_toolkit.plot_info import PlotType
from data_processing_toolkit.excel_formulas import (
    convert_to_excel_address,
    get_cell_value,
)
from data_processing_toolkit.file_handler import (
    get_file,
    get_path_to_file,
    create_folder_if_not_exists,
)


def generate_plots(
    plot_infos,
    primary_marker="o",
    secondary_marker="o",
    secondary_marker_2="s",
    secondary_marker_3="s",
    fill_primary_marker=True,
    error_bar_cap_size=3,
    error_bar_marker="none",
    dashed_line_style = "dashed",
    dashed_line_color = "k",
    fit_line_style = "dashed",
    fit_line_color = "#ff6666",
    color_scheme=[
        "#000000",
        "#ff0000",
        "#1f77b4",
        "#ff7f0e",
        "#2ca02c",
        "#9467bd",
        "#8c564b",
        "#e377c2",
        "#7f7f7f",
        "#bcbd22",
        "#17becf",
    ],
):
    """
    Module conatining a function for plotting data from a selected excel file

    Parameters
    ----------
    plot_infos : list[PlotInfo]
        formatting information for each plot that is to be generated
    primary_marker : str, optional
        if plotting a scatter plot, this is the marker style (see matplotlib documentation)
        of the primary series, by default "o"
    secondary_marker : str, optional
        if plotting a scatter plot, this is the marker style (see matplotlib documentation)
        of the 1st secondary series, by default "o"
    secondary_marker_2 : str, optional
        if plotting a scatter plot, this is the marker style (see matplotlib documentation)
        of the 2nd secondary series, by default "s"
    secondary_marker_3 : str, optional
        if plotting a scatter plot, this is the marker style (see matplotlib documentation)
        of the 3rd secondary series, by default "s"
    fill_primary_marker : bool, optional
        if plotting a scatter plot, primary marker will be filled if True, and secondary
        markers will not be, by default True
    error_bar_cap_size : int, optional
        size of the error bar cap, by default 3
    error_bar_marker : str, optional
        if plotting a scatter plot, this is the marker style (see matplotlib documentation)
        of the error bar, t, by default "none"
    dashed_line_style : str, optional
        line style (see matplotlib documentation) of all dashed lines added to the figure,
        by default "dashed"
    dashed_line_color : str, optional
        line color (see matplotlib documentation) of all dashed lines added to the figure,
        by default "k"
    fit_line_style : str, optional
        line style (see matplotlib documentation) of all fit lines added to the figure,
        by default "dashed
    fit_line_color : str, optional
        line color (see matplotlib documentation) of all dashed lines added to the figure,
        by default "#ff6666"
    color_scheme : list[str], optional
        color_scheme[i] represents the color used in plotting the ith series , 
        by default [ "#000000", "#ff0000", "#1f77b4", "#ff7f0e", "#2ca02c", "#9467bd", "#8c564b",
            "#e377c2", "#7f7f7f", "#bcbd22", "#17becf", ]
    """
    # MATPLOTLIB CONSTANTS, CAN BE EDITED TO SUIT YOUR PREFERENCES

    # Scatter plot markers
    primary_face_color = "none" if not fill_primary_marker else None
    secondary_face_color = "none" if fill_primary_marker else None

    excel_file_path = get_file()  # select excel file to generate plots from
    if excel_file_path:
        # load excel workbook
        root_folder = get_path_to_file(excel_file_path)
        workbook = load_workbook(excel_file_path, data_only=True)
        sheetnames = workbook.sheetnames
        print(f"Excel file selected: {excel_file_path}")

        for sheetname in sheetnames:
            if sheetname in plot_infos:
                # load sheet, sheet formatting, and relevant sheet content
                plot_info = plot_infos[sheetname]
                sheet = workbook[sheetname]
                row_count = sheet.max_row
                col_count = sheet.max_column
                series_count = 0
                print(f"\tProcessing sheet: {sheetname}")

                # locate all target column numbers based on headers
                # NOTE - header constants are case-sensitive and must be entirely lower-case
                series_name_col = None
                x_col = y_col = None
                y_pos_error_col = y_neg_error_col = None
                x_pos_error_col = x_neg_error_col = None
                for col_no in range(col_count):
                    col_header = sheet[
                        convert_to_excel_address(row_no=0, col_no=col_no)
                    ].value
                    if (
                        col_header == constants.HEADER_HEADER
                        and series_name_col is None
                    ):
                        series_name_col = col_no
                    elif col_header == constants.X_HEADER and x_col is None:
                        x_col = col_no
                    elif col_header == constants.Y_HEADER and y_col is None:
                        y_col = col_no
                    elif col_header == constants.Y_ERROR_HEADER:
                        y_pos_error_col = y_neg_error_col = col_no
                    elif (
                        col_header == constants.Y_POS_ERROR_HEADER
                        and y_pos_error_col is None
                    ):
                        y_pos_error_col = col_no
                    elif (
                        col_header == constants.Y_NEG_ERROR_HEADER
                        and y_neg_error_col is None
                    ):
                        y_neg_error_col = col_no
                    elif col_header == constants.X_ERROR_HEADER:
                        x_pos_error_col = x_neg_error_col = col_no
                    elif (
                        col_header == constants.X_POS_ERROR_HEADER
                        and x_pos_error_col is None
                    ):
                        x_pos_error_col = col_no
                    elif (
                        col_header == constants.X_NEG_ERROR_HEADER
                        and x_neg_error_col is None
                    ):
                        x_neg_error_col = col_no

                    target_cols = (
                        series_name_col,
                        x_col,
                        y_col,
                        y_pos_error_col,
                        y_neg_error_col,
                        x_pos_error_col,
                        x_neg_error_col,
                    )
                    if None not in target_cols:
                        break

                # load contents of target columns while keeping track of the
                # x and y extrema of the data, as well as the number of data
                # series to be plotted
                plot_data = {}
                min_y = max_y = None
                min_x = max_x = None
                for row_no in range(1, row_count):
                    (
                        series_name,
                        x_val,
                        y_val,
                        y_pos_error,
                        y_neg_error,
                        x_pos_error,
                        x_neg_error,
                    ) = tuple(
                        map(
                            lambda col_no: get_cell_value(
                                sheet,
                                row_no,
                                col_no,
                                is_number=col_no is not None
                                and series_name_col is not col_no,
                            ),
                            target_cols,
                        )
                    )
                    series_name = "blank" if series_name is None else series_name
                    if None not in (x_val, y_val):
                        min_y = min(min_y, y_val) if min_y is not None else y_val
                        max_y = max(max_y, y_val) if max_y is not None else y_val
                        min_x = min(min_x, x_val) if min_x is not None else x_val
                        max_x = max(max_x, x_val) if max_x is not None else x_val
                        series_data = plot_data.get(series_name, {})
                        if (
                            constants.INDEX_HEADER not in series_data
                            and (
                                not str(series_name).startswith(
                                    constants.SECONDARY_SERIES_PREFIX
                                )
                            )
                            and (
                                not str(series_name).startswith(
                                    constants.SECONDARY_2_SERIES_PREFIX
                                )
                            )
                            and (
                                not str(series_name).startswith(
                                    constants.SECONDARY_3_SERIES_PREFIX
                                )
                            )
                        ):
                            series_data[constants.INDEX_HEADER] = series_count
                            series_count += 1

                        series_data[constants.X_HEADER] = series_data.get(
                            constants.X_HEADER, []
                        ) + [x_val]
                        series_data[constants.Y_HEADER] = series_data.get(
                            constants.Y_HEADER, []
                        ) + [y_val]
                        series_data[constants.Y_POS_ERROR_HEADER] = series_data.get(
                            constants.Y_POS_ERROR_HEADER, []
                        ) + [y_pos_error if y_pos_error is not None else 0]
                        series_data[constants.Y_NEG_ERROR_HEADER] = series_data.get(
                            constants.Y_NEG_ERROR_HEADER, []
                        ) + [y_neg_error if y_neg_error is not None else 0]
                        series_data[constants.X_POS_ERROR_HEADER] = series_data.get(
                            constants.X_POS_ERROR_HEADER, []
                        ) + [x_pos_error if x_pos_error is not None else 0]
                        series_data[constants.X_NEG_ERROR_HEADER] = series_data.get(
                            constants.X_NEG_ERROR_HEADER, []
                        ) + [x_neg_error if x_neg_error is not None else 0]
                        plot_data[series_name] = series_data

                # create matplotlib figure or subplot (depending on PlotInfo formatting)
                # NOTE - for proper functionality, all excel sheets which are part of the same
                # subplot figure must be side by side in the excel workbook. The order they occur
                # in the excel sheet are irrelevant (as long as they are side-by-side) since their
                # order is defined in the PlotInfo list.
                if plot_info.suplot_tuple is not None:
                    (
                        subplot_row_count,
                        subplot_col_count,
                        subplot_index,
                    ) = plot_info.suplot_tuple
                    if subplot_index == 1:
                        plt.figure(sheetnames.index(sheetname), dpi=plot_info.dpi)
                    plt.subplot(subplot_row_count, subplot_col_count, subplot_index)
                else:
                    plt.figure(sheetnames.index(sheetname), dpi=plot_info.dpi)

                # sort all series names, pushing secondary series to the right end of
                # the list
                sorted_series_names = list(
                    map(
                        lambda key: [
                            key,
                            plot_data[key].get(constants.INDEX_HEADER, None),
                        ],
                        plot_data.keys(),
                    )
                )
                sorted_series_names.sort(key=lambda x: (x[1] is None, x[1]))
                if plot_info.reverse_series:
                    sorted_series_names.reverse()

                # PLOT FIGURE AND IMPLEMENT SPECIFIED FORMATTING
                for series_name, series_idx in sorted_series_names:
                    # if current series is a secondary series, remove legend
                    # and provide same color scheme as original series
                    is_secondary_series = series_idx is None
                    secondary_idx = (
                        0
                        if not is_secondary_series
                        else 1
                        if (
                            str(series_name).startswith(
                                constants.SECONDARY_SERIES_PREFIX
                            )
                        )
                        else 2
                        if (
                            str(series_name).startswith(
                                constants.SECONDARY_2_SERIES_PREFIX
                            )
                        )
                        else 3
                    )

                    true_idx = series_idx
                    legend = (
                        "_nolegend_"
                        if is_secondary_series or not plot_info.legends
                        else plot_info.legends[true_idx]
                    )
                    if is_secondary_series:
                        true_name = series_name
                        for sec_keyword in (
                            constants.SECONDARY_SERIES_PREFIX,
                            constants.SECONDARY_2_SERIES_PREFIX,
                            constants.SECONDARY_3_SERIES_PREFIX,
                        ):
                            if true_name.startswith(sec_keyword):
                                true_name = true_name[len(sec_keyword) :]
                                break
                        for sub_series_name, sub_idx in sorted_series_names:
                            if sub_series_name == true_name:
                                true_idx = sub_idx
                                break
                        if true_idx is None:
                            true_idx = series_count
                            series_count += 1
                    series_color = (
                        plot_info.color_map[true_idx]
                        if plot_info.color_map
                        else color_scheme[true_idx]
                    )

                    # error bar values
                    y_pos_errors, y_neg_errors = (
                        plot_data[series_name][constants.Y_POS_ERROR_HEADER],
                        plot_data[series_name][constants.Y_NEG_ERROR_HEADER],
                    )
                    x_pos_errors, x_neg_errors = (
                        plot_data[series_name][constants.X_POS_ERROR_HEADER],
                        plot_data[series_name][constants.X_NEG_ERROR_HEADER],
                    )
                    has_y_error = max(*y_pos_errors, *y_neg_errors) != 0
                    has_x_error = max(*x_pos_errors, *x_neg_errors) != 0

                    # plot scatter/line plot (based on PlotInfo) with error bars
                    if plot_info.plot_type == PlotType.SCATTER:
                        x, y = (
                            plot_data[series_name][constants.X_HEADER],
                            plot_data[series_name][constants.Y_HEADER],
                        )
                        x = np.array(x) + (
                            true_idx * plot_info.secondary_series_stagger
                        )  # secondary data points can be shifted horizontally
                        # by a set offset for clarity

                        marker = (
                            primary_marker
                            if not is_secondary_series
                            else secondary_marker
                            if secondary_idx == 1
                            else secondary_marker_2
                            if secondary_idx == 2
                            else secondary_marker_3
                        )
                        face_color = (
                            primary_face_color
                            if (not is_secondary_series)
                            else secondary_face_color
                            if (secondary_idx == 1)
                            else primary_face_color
                            if secondary_idx == 2
                            else secondary_face_color
                        )
                        plt.scatter(
                            x,
                            y,
                            color=series_color,
                            facecolors=face_color,
                            marker=marker,
                            label=legend,
                            s=plot_info.marker_size,
                        )

                        if has_y_error or has_x_error:
                            y_error = (
                                np.array(list(zip(y_neg_errors, y_pos_errors))).T
                                if has_y_error
                                else None
                            )
                            x_error = (
                                np.array(list(zip(x_neg_errors, x_pos_errors))).T
                                if has_x_error
                                else None
                            )
                            plt.errorbar(
                                x,
                                y,
                                yerr=y_error,
                                xerr=x_error,
                                fmt=error_bar_marker,
                                error_bar_cap_size=error_bar_cap_size,
                                ecolor=series_color,
                                label="_nolegend_",
                            )
                    elif plot_info.plot_type == PlotType.LINE:
                        x, y = (
                            plot_data[series_name][constants.X_HEADER],
                            plot_data[series_name][constants.Y_HEADER],
                        )
                        if plot_info.sort_data:
                            x = np.array(x)
                            y = np.array(y)
                            sorting = np.argsort(x)
                            x = x[sorting]
                            y = y[sorting]

                        plt.plot(x, y, color=series_color, label=legend)

                        if has_y_error:
                            plt.fill_between(
                                x,
                                np.array(y) - np.array(y_neg_errors),
                                np.array(y) + np.array(y_pos_errors),
                                color=series_color,
                                alpha=0.25,
                                label="_nolegend_",
                            )
                        if has_x_error:
                            plt.fill_between(
                                y,
                                np.array(x) - np.array(x_neg_errors),
                                np.array(x) + np.array(x_pos_errors),
                                color=series_color,
                                alpha=0.25,
                                label="_nolegend_",
                            )

                    if plot_info.draw_line_at_maximum:
                        # draw a vertical line that passes through the
                        # maximum y-value of the series
                        if len(x) > 0:
                            y_min, y_max = np.min(y), np.max(y)
                            x_max = x[np.nanargmax(y)]
                            if plot_info.draw_line_at_maximum:
                                plt.vlines(
                                    x_max,
                                    min(0, y_min),
                                    max(0, y_max),
                                    linestyles="dashed",
                                    label="_nolegend_",
                                )

                for i, fit_line in enumerate(plot_info.fit_lines):
                    # draw lines on the figure defined by y = fit_function(x)
                    # with a domain of [fit_start_x, fit_end_x]
                    fit_start_x, fit_end_x, fit_function = fit_line
                    fit_x = np.linspace(fit_start_x, fit_end_x)
                    fit_y = list(map(fit_function, fit_x))
                    plt.plot(
                        fit_x,
                        fit_y,
                        color=fit_line_color
                        if not plot_info.use_color_map_for_fit_color
                        else color_scheme[i],
                        linestyle=fit_line_style,
                    )

                if plot_info.legends:
                    # include series legends based on specified formatting
                    if plot_info.reverse_legends:
                        handles, labels = plt.gca().get_legend_handles_labels()
                        plt.gca().legend(
                            handles[::-1],
                            labels[::-1],
                            fontsize=plot_info.legend_size,
                            loc=plot_info.legend_loc,
                            ncol=plot_info.legend_n_cols,
                            bbox_to_anchor=plot_info.legend_bbox_anchor,
                        )
                    else:
                        plt.legend(
                            fontsize=plot_info.legend_size,
                            loc=plot_info.legend_loc,
                            ncol=plot_info.legend_n_cols,
                            bbox_to_anchor=plot_info.legend_bbox_anchor,
                        )

                if plot_info.vline_at_x is not None:
                    # draw vertical line that passes through a specified x-value
                    plt.vlines(
                        plot_info.vline_at_x,
                        0 if np.sign(min_y) == np.sign(max_y) else min_y,
                        max_y if max_y > 0 else min_y,
                        linestyles=dashed_line_style,
                        color_scheme=dashed_line_color,
                        label="_nolegend_",
                    )
                if plot_info.hline_at_y is not None:
                    # draw horizontal line that passes through a specified y-value
                    plt.hlines(
                        plot_info.hline_at_y,
                        0 if np.sign(min_x) == np.sign(max_x) else min_x,
                        max_x + (series_count * plot_info.secondary_series_stagger)
                        if max_x > 0
                        else min_x,
                        linestyles=dashed_line_style,
                        color_scheme=dashed_line_color,
                        label="_nolegend_",
                    )

                # convert axes to logarithm scale based on specified formatting
                if plot_info.yaxis_log_scale:
                    plt.yscale("log", base=10)
                if plot_info.xaxis_log_scale:
                    plt.xscale("log", base=10)

                # additional figure formatting
                plt.ylim(plot_info.ylim)
                plt.xlim(plot_info.xlim)
                plt.title(plot_info.title, loc="left")
                plt.xlabel(plot_info.xlabel, fontsize=plot_info.font_size)
                plt.ylabel(plot_info.ylabel, fontsize=plot_info.font_size)
                plt.xticks(fontsize=plot_info.font_size)
                plt.yticks(fontsize=plot_info.font_size)
                plt.gca().tick_params(axis="both", which="both", direction="in")
                if not plot_info.show_x_axis_numbers:
                    plt.gca().axes.xaxis.set_ticklabels([])
                if not plot_info.show_y_axis_numbers:
                    plt.gca().axes.yaxis.set_ticklabels([])
                if plot_info.figure_size_in_inches:
                    plt.gcf().set_size_inches(*plot_info.figure_size_in_inches)

                # save figure using specified file extension
                if (
                    plot_info.suplot_tuple is None
                    or (plot_info.suplot_tuple[0] * plot_info.suplot_tuple[1])
                    <= plot_info.suplot_tuple[2]
                ):
                    save_folder = create_folder_if_not_exists(
                        join(root_folder, plot_info.save_sub_directory)
                    )
                    save_destination = join(
                        save_folder, sheetname + plot_info.save_file_extension
                    )
                    plt.savefig(save_destination)
                    plt.close()
