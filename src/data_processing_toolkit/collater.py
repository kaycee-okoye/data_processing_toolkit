"""
    Module conatins a function for combining excel files into a single excel file.
    The function assumes that the excel files being combined have similar structures.
    It looks through each sheet in each file for target column headers (e.g. 'Average') 
    then writes the content of each target column into the corresponding sheet in the 
    combined excel file.
"""

from os.path import join
from os import startfile
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series, marker
import data_processing_toolkit.constants as constants
from data_processing_toolkit.excel_formulas import (
    convert_to_excel_address,
    create_sheet_name,
    get_cell_value,
    insert_openpyxl_chart,
)
from data_processing_toolkit.file_handler import (
    get_excel_files_from_folder,
    get_filename,
)

def insert_point_analysis_chart(
    pt_analysis_sheet, title: str, n_files: int, sheet_index: int, x_axis_label:str
):
    """
    Function to insert dynamic plots into point analysis worksheet.

    Parameters
    ----------
    pt_analysis_sheet : xlsxwriter worksheet
        point analysis worksheet in destination workbook
    title : str
        title of the chart being inserted in the destination workbook
    n_files : int
        number of excel files from which data is being copied
    sheet_index : int
        index of the copied worksheet in the destination workbook
    x_axis_label : str
        title of x-axis
    """

    chart = ScatterChart()
    size = Reference(worksheet=pt_analysis_sheet, min_col=2, min_row=2, max_row=10)
    xvalue = Reference(
        worksheet=pt_analysis_sheet, min_col=2, max_col=n_files + 1, min_row=2
    )
    y_value = Reference(
        worksheet=pt_analysis_sheet,
        min_col=2,
        max_col=n_files + 1,
        min_row=2 + sheet_index,
    )
    series = Series(values=y_value, xvalues=xvalue, zvalues=size, title=title)
    series.marker = marker.Marker("circle")
    series.graphicalProperties.line.noFill = True

    chart.append(series)
    chart.y_axis.title = title
    chart.x_axis.title = x_axis_label
    chart.y_axis.majorGridlines = None
    chart.title = title
    pt_analysis_sheet.add_chart(
        chart, convert_to_excel_address(row_no=2, col_no=n_files)
    )


def insert_point_analysis_formula(
    pt_analysis_sheet,
    sheet_index: int,
    sheet_title: str,
    file_index_: int,
    space_between_targets: int,
    target_headers
):
    """
    Function to insert point analysis sheet formulas for all target columns
    in a given worksheet of a given file being combined.

    Parameters
    ----------
    pt_analysis_sheet : xlsxwriter worksheet
        point analysis worksheet in destination workbook
    sheet_index : int
        index of the copied worksheet in the destination workbook
    sheet_title : str
        title of the copied worksheet in the destination workbook
    file_index_ : int
        index of excel file from which data is being copied
    space_between_targets : int
        no. of columns placed between the columns of
        target columns of a given source excel file in the combined excel file
    target_headers : list[str]
        for each sheet in each excel file being combined, columns with header values
        in this list will be added to the combined excel file
    """
    point_address = convert_to_excel_address(0, 1 + file_index_)
    for header_idx, target_header_ in enumerate(target_headers):
        pt_analysis_sheet[
            convert_to_excel_address(
                sheet_index, +(space_between_targets * header_idx)
            )
        ] = f"{sheet_title} ({target_header_})"
        pt_analysis_sheet[
            convert_to_excel_address(sheet_index, 1 + file_index_)
        ] = f'=INDIRECT(ADDRESS({point_address}+1, {file_index_ +(space_between_targets*header_idx)+ 1},,,"{sheet_title}"))'

def collate_excel_files(root_folder, target_headers, independent_variable_sheetname=""):
    """
    Function for combining excel files into a single excel file.

    Parameters
    ----------
    root_folder : str
        folder containing excel files to be combined
    target_headers : list[str]
        for each sheet in each excel file being combined, columns with header values
        in this list will be added to the combined excel file
    independent_variable_sheetname : str, optional
        sheet name of data set which all other data sets will be plotted against. If not
        provided, no plots will be added in the combined excel file, by default ""

    NOTE
    When complete, the combined excel file will be automatically opened
    """
    target_headers_set = set(target_headers)  # used to lookup target headers in O(1) time
    # source excel file
    if root_folder:
        print(f"Folder selected: {root_folder}")
        # get excel files from the root_folder
        excel_file_paths = get_excel_files_from_folder(root_folder)
        excel_filenames = [get_filename(file_path) for file_path in excel_file_paths]
        n_files = len(excel_file_paths)
        rel_space_between_targets = (
            n_files + 2
        )  # no. of columns placed between the columns of
        # target columns of a given source excel file in the combined excel file

        if n_files:
            dst_workbook = Workbook()
            # create sheet for storing point analysis data which allows you view values
            # of each files at a specific value of the independent variable
            pt_analysis_sheet = dst_workbook.create_sheet(
                create_sheet_name(constants.POINT_ANALYSIS_LABEL)
            )
            pt_analysis_sheet[convert_to_excel_address(0, 0)] = "Column Index ->"
            max_row_count = 0  # maximum number of rows across files

            for file_index in range(n_files):
                # load source workbook
                src_workbook = load_workbook(excel_file_paths[file_index], data_only=True)
                sheetnames = src_workbook.sheetnames
                excel_filename = excel_filenames[file_index]
                print(f"Reading excel file: {excel_filename}")

                for idx, target_header in enumerate(target_headers):
                    # add filename for each target header to the point analysis sheet
                    # to indicate the row where it's data will be displayed
                    pt_analysis_sheet[
                        convert_to_excel_address(1, 1 + (rel_space_between_targets * idx))
                    ] = f"{excel_filename} ({target_header})"

                for sheetname in sheetnames:
                    # loop through each sheet to extract target columns and write
                    # to the combined excel file
                    sheet = src_workbook[sheetname]
                    row_count = sheet.max_row
                    col_count = sheet.max_colum
                    max_row_count = max(row_count, max_row_count)
                    contains_a_target_column = False

                    for col in range(col_count):
                        # check if column's header matches any of the target headers
                        header_cell_value = sheet[convert_to_excel_address(0, col)].value
                        if header_cell_value in target_headers_set:
                            if sheetname in dst_workbook.sheetnames:
                                dst_sheet = dst_workbook[sheetname]
                            else:
                                dst_sheet = dst_workbook.create_sheet(sheetname)
                            sheet_index = dst_workbook.sheetnames.index(sheetname)
                            if ~contains_a_target_column:
                                # add formulas for this sheet into the point analysis sheet
                                # if this hasn't been done before
                                contains_a_target_column = True
                                insert_point_analysis_formula(
                                    pt_analysis_sheet,
                                    sheet_index,
                                    sheetname,
                                    file_index,
                                    rel_space_between_targets,
                                    target_headers
                                )

                            # write filename as header in the destination worksheet
                            header_index = target_headers.index(header_cell_value)
                            dst_sheet[
                                convert_to_excel_address(
                                    0,
                                    file_index + (rel_space_between_targets * header_index),
                                )
                            ] = f"{excel_filename} ({header_cell_value})"

                            for row in range(1, row_count + 1):
                                # transfer target column to destination worksheet
                                print(
                                    f"\tTransferring to destination worksheet: {header_cell_value}"
                                )
                                cell_content = get_cell_value(
                                    sheet,
                                    row,
                                    col,
                                    is_number=True,
                                    default_value=constants.ERROR_VALUE_FOR_INPUT,
                                )
                                dst_sheet[
                                    convert_to_excel_address(
                                        row,
                                        file_index
                                        + (rel_space_between_targets * header_index),
                                    )
                                ] = cell_content

            # save combined excel workbook to root folder
            destination_path = join(root_folder, f"-{constants.COLLATER_EXCEL_FILENAME}")
            dst_workbook.save(destination_path)
            dst_workbook = load_workbook(destination_path)
            dst_workbook.active = dst_workbook.worksheets[1]
            if constants.EXCEL_DEFAULT_SHEETNAME in dst_workbook.sheetnames:
                # remove empty excel sheet that's usually added by default
                dst_workbook.remove(dst_workbook.worksheets[0])

            if independent_variable_sheetname:
                # add plots to all sheets in the desitination workbook,
                # plotting the data of each column against the
                # dependent variable. Do this here, after saving the worksheet once, otherwise
                # the plots will come up empty
                pt_analysis_sheet = dst_workbook.create_sheet(
                    create_sheet_name(constants.POINT_ANALYSIS_LABEL)
                )
                x_sheet = dst_workbook[independent_variable_sheetname]
                dst_sheetnames = dst_workbook.sheetnames
                for sheetname in dst_sheetnames:
                    if sheetname != constants.POINT_ANALYSIS_LABEL:
                        for i, target_header in enumerate(target_headers):
                            cols = [
                                1
                                for i in range(
                                    (rel_space_between_targets * i),
                                    n_files + (rel_space_between_targets * i),
                                )
                            ]
                            min_rows = [1 for i in range(n_files)]
                            max_rows = [max_row_count for i in range(n_files)]
                            insert_openpyxl_chart(
                                output_sheet=dst_workbook[sheetname],
                                title=sheetname,
                                x_title=independent_variable_sheetname,
                                y_title=sheetname,
                                x_sheet=x_sheet,
                                x_cols=cols,
                                x_min_rows=min_rows,
                                x_max_rows=max_rows,
                                y_sheet=dst_workbook[sheetname],
                                y_cols=cols,
                                y_min_rows=min_rows,
                                y_max_rows=max_rows,
                                labels=excel_filenames,
                                location=convert_to_excel_address(
                                    0, n_files + (rel_space_between_targets * i)
                                ),
                            )
                    else:
                        # insert dynamic plots into point analysis worksheet
                        for sheet_index, sheetname in enumerate(dst_sheetnames):
                            if sheetname != create_sheet_name(
                                constants.POINT_ANALYSIS_LABEL
                            ):
                                insert_point_analysis_chart(
                                    pt_analysis_sheet,
                                    sheetname,
                                    n_files,
                                    sheet_index,
                                    target_headers
                                )

            # save the file once more
            dst_workbook.active = 0  # make sure the first worksheet is open
            # when launching
            dst_workbook.save(destination_path)
            startfile(destination_path)  # open combined excel file
