"""
    Module conatining a script for plotting data from a selected excel file.
    A list of PlotInfo objects corresponding to the desired sheetnames
    are used to provide formatting information for the figures.

    NOTE
    - PLOTS variable in plot_info.py will need to be populated in order to use plotter.py
    - the input source for plot data is from an excel file. However, this can easily be
        modified for the users specific use case.

"""

from os.path import join
from openpyxl import load_workbook
from matplotlib import pyplot as plt
import numpy as np
import constants
from plot_info import PlotType, PLOTS
from excel_formulas import convert_to_excel_address, get_cell_value
from file_handler import get_file, get_path_to_file, create_folder_if_not_exists

# MATPLOTLIB CONSTANTS, CAN BE EDITED TO SUIT YOUR PREFERENCES

# color scheme used for each data series
COLOR_SCHEME = [
    "#000000",
    "#ff0000",
    "#1f77b4",
    "#ff7f0e",
    "#2ca02c",
    "#9467bd",
    "#8c564b",
    "#e377c2",
    "#7f7f7f",
    "#bcbd22",
    "#17becf",
]
# Scatter plot markers
FILL_PRIMARY_MARKER = True
PRIMARY_FACE_COLOR = "none" if not FILL_PRIMARY_MARKER else None
SECONDARY_FACE_COLOR = "none" if FILL_PRIMARY_MARKER else None
PRIMARY_MARKER = "o"
SECONDARY_MARKER = "o"
SECONDARY_2_MARKER = "s"
SECONDARY_3_MARKER = "s"

# Error Bars
CAP_SIZE = 3
ERROR_MARKER = "none"

# Dashed Lines
DASHED_LINE_STYLE = "dashed"
DASHED_LINE_COLOR = "k"

# Fit Lines
FIT_LINE_STYLE = "dashed"
FIT_LINE_COLOR = "#ff6666"

excel_file_path = get_file()  # select excel file to generate plots from
if excel_file_path:
    # load excel workbook
    root_folder = get_path_to_file(excel_file_path)
    workbook = load_workbook(excel_file_path, data_only=True)
    sheetnames = workbook.sheetnames
    print(f"Excel file selected: {excel_file_path}")

    for sheetname in sheetnames:
        if sheetname in PLOTS:
            # load sheet, sheet formatting, and relevant sheet content
            plot_info = PLOTS[sheetname]
            sheet = workbook[sheetname]
            row_count = sheet.max_row
            col_count = sheet.max_column
            SERIES_COUNT = 0
            print(f"\tProcessing sheet: {sheetname}")

            # locate all target column numbers based on headers
            # NOTE - header constants are case-sensitive and must be entirely lower-case
            SERIES_NAME_COL = None
            X_COL = Y_COL = None
            Y_POS_ERRROR_COL = Y_NEG_ERROR_COL = None
            X_POS_ERROR_COL = X_NEG_ERROR_COL = None
            for col_no in range(col_count):
                col_header = sheet[
                    convert_to_excel_address(row_no=0, col_no=col_no)
                ].value
                if col_header == constants.HEADER_HEADER and SERIES_NAME_COL is None:
                    SERIES_NAME_COL = col_no
                elif col_header == constants.X_HEADER and X_COL is None:
                    X_COL = col_no
                elif col_header == constants.Y_HEADER and Y_COL is None:
                    Y_COL = col_no
                elif col_header == constants.Y_ERROR_HEADER:
                    Y_POS_ERRROR_COL = Y_NEG_ERROR_COL = col_no
                elif (
                    col_header == constants.Y_POS_ERROR_HEADER
                    and Y_POS_ERRROR_COL is None
                ):
                    Y_POS_ERRROR_COL = col_no
                elif (
                    col_header == constants.Y_NEG_ERROR_HEADER
                    and Y_NEG_ERROR_COL is None
                ):
                    Y_NEG_ERROR_COL = col_no
                elif col_header == constants.X_ERROR_HEADER:
                    X_POS_ERROR_COL = X_NEG_ERROR_COL = col_no
                elif (
                    col_header == constants.X_POS_ERROR_HEADER
                    and X_POS_ERROR_COL is None
                ):
                    X_POS_ERROR_COL = col_no
                elif (
                    col_header == constants.X_NEG_ERROR_HEADER
                    and X_NEG_ERROR_COL is None
                ):
                    X_NEG_ERROR_COL = col_no

                target_cols = (
                    SERIES_NAME_COL,
                    X_COL,
                    Y_COL,
                    Y_POS_ERRROR_COL,
                    Y_NEG_ERROR_COL,
                    X_POS_ERROR_COL,
                    X_NEG_ERROR_COL,
                )
                if None not in target_cols:
                    break

            # load contents of target columns while keeping track of the
            # x and y extrema of the data, as well as the number of data
            # series to be plotted
            plot_data = {}
            MIN_Y = MAX_Y = None
            MIN_X = MAX_X = None
            for row_no in range(1, row_count):
                (
                    series_name,
                    x_val,
                    y_val,
                    y_pos_error,
                    y_neg_error,
                    x_pos_error,
                    x_neg_error,
                ) = tuple(
                    map(
                        lambda col_no: get_cell_value(
                            sheet,
                            row_no,
                            col_no,
                            is_number=col_no is not None and SERIES_NAME_COL is not col_no,
                        ),
                        target_cols,
                    )
                )
                series_name = "blank" if series_name is None else series_name
                if None not in (x_val, y_val):
                    MIN_Y = min(MIN_Y, y_val) if MIN_Y is not None else y_val
                    MAX_Y = max(MAX_Y, y_val) if MAX_Y is not None else y_val
                    MIN_X = min(MIN_X, x_val) if MIN_X is not None else x_val
                    MAX_X = max(MAX_X, x_val) if MAX_X is not None else x_val
                    series_data = plot_data.get(series_name, {})
                    if (
                        constants.INDEX_HEADER not in series_data
                        and (
                            not str(series_name).startswith(
                                constants.SECONDARY_SERIES_PREFIX
                            )
                        )
                        and (
                            not str(series_name).startswith(
                                constants.SECONDARY_2_SERIES_PREFIX
                            )
                        )
                        and (
                            not str(series_name).startswith(
                                constants.SECONDARY_3_SERIES_PREFIX
                            )
                        )
                    ):
                        series_data[constants.INDEX_HEADER] = SERIES_COUNT
                        SERIES_COUNT += 1

                    series_data[constants.X_HEADER] = series_data.get(
                        constants.X_HEADER, []
                    ) + [x_val]
                    series_data[constants.Y_HEADER] = series_data.get(
                        constants.Y_HEADER, []
                    ) + [y_val]
                    series_data[constants.Y_POS_ERROR_HEADER] = series_data.get(
                        constants.Y_POS_ERROR_HEADER, []
                    ) + [y_pos_error if y_pos_error is not None else 0]
                    series_data[constants.Y_NEG_ERROR_HEADER] = series_data.get(
                        constants.Y_NEG_ERROR_HEADER, []
                    ) + [y_neg_error if y_neg_error is not None else 0]
                    series_data[constants.X_POS_ERROR_HEADER] = series_data.get(
                        constants.X_POS_ERROR_HEADER, []
                    ) + [x_pos_error if x_pos_error is not None else 0]
                    series_data[constants.X_NEG_ERROR_HEADER] = series_data.get(
                        constants.X_NEG_ERROR_HEADER, []
                    ) + [x_neg_error if x_neg_error is not None else 0]
                    plot_data[series_name] = series_data

            # create matplotlib figure or subplot (depending on PlotInfo formatting)
            # NOTE - for proper functionality, all excel sheets which are part of the same
            # subplot figure must be side by side in the excel workbook. The order they occur
            # in the excel sheet are irrelevant (as long as they are side-by-side) since their
            # order is defined in the PlotInfo list.
            if plot_info.suplot_tuple is not None:
                (
                    subplot_row_count,
                    subplot_col_count,
                    subplot_index,
                ) = plot_info.suplot_tuple
                if subplot_index == 1:
                    plt.figure(sheetnames.index(sheetname), dpi=plot_info.dpi)
                plt.subplot(subplot_row_count, subplot_col_count, subplot_index)
            else:
                plt.figure(sheetnames.index(sheetname), dpi=plot_info.dpi)

            # sort all series names, pushing secondary series to the right end of
            # the list
            sorted_series_names = list(
                map(
                    lambda key: [key, plot_data[key].get(constants.INDEX_HEADER, None)],
                    plot_data.keys(),
                )
            )
            sorted_series_names.sort(key=lambda x: (x[1] is None, x[1]))
            if plot_info.reverse_series:
                sorted_series_names.reverse()

            # PLOT FIGURE AND IMPLEMENT SPECIFIED FORMATTING
            for series_name, series_idx in sorted_series_names:
                # if current series is a secondary series, remove legend
                # and provide same color scheme as original series
                is_secondary_series = series_idx is None
                SECONDARY_IDX = (
                    0
                    if not is_secondary_series
                    else 1
                    if (str(series_name).startswith(constants.SECONDARY_SERIES_PREFIX))
                    else 2
                    if (
                        str(series_name).startswith(constants.SECONDARY_2_SERIES_PREFIX)
                    )
                    else 3
                )

                TRUE_IDX = series_idx
                legend = (
                    "_nolegend_"
                    if is_secondary_series or not plot_info.legends
                    else plot_info.legends[TRUE_IDX]
                )
                if is_secondary_series:
                    true_name = series_name
                    for sec_keyword in (
                        constants.SECONDARY_SERIES_PREFIX,
                        constants.SECONDARY_2_SERIES_PREFIX,
                        constants.SECONDARY_3_SERIES_PREFIX,
                    ):
                        if true_name.startswith(sec_keyword):
                            true_name = true_name[len(sec_keyword) :]
                            break
                    for sub_series_name, sub_idx in sorted_series_names:
                        if sub_series_name == true_name:
                            TRUE_IDX = sub_idx
                            break
                    if TRUE_IDX is None:
                        TRUE_IDX = SERIES_COUNT
                        SERIES_COUNT += 1
                series_color = (
                    plot_info.color_map[TRUE_IDX]
                    if plot_info.color_map
                    else COLOR_SCHEME[TRUE_IDX]
                )

                # error bar values
                y_pos_errors, y_neg_errors = (
                    plot_data[series_name][constants.Y_POS_ERROR_HEADER],
                    plot_data[series_name][constants.Y_NEG_ERROR_HEADER],
                )
                x_pos_errors, x_neg_errors = (
                    plot_data[series_name][constants.X_POS_ERROR_HEADER],
                    plot_data[series_name][constants.X_NEG_ERROR_HEADER],
                )
                has_y_error = max(*y_pos_errors, *y_neg_errors) != 0
                has_x_error = max(*x_pos_errors, *x_neg_errors) != 0

                # plot scatter/line plot (based on PlotInfo) with error bars
                if plot_info.plot_type == PlotType.SCATTER:
                    x, y = (
                        plot_data[series_name][constants.X_HEADER],
                        plot_data[series_name][constants.Y_HEADER],
                    )
                    x = np.array(x) + (
                        TRUE_IDX * plot_info.secondary_series_stagger
                    )  # secondary data points can be shifted horizontally
                    # by a set offset for clarity

                    MARKER = (
                        PRIMARY_MARKER
                        if not is_secondary_series
                        else SECONDARY_MARKER
                        if SECONDARY_IDX == 1
                        else SECONDARY_2_MARKER
                        if SECONDARY_IDX == 2
                        else SECONDARY_3_MARKER
                    )
                    FACE_COLOR = (
                        PRIMARY_FACE_COLOR
                        if (not is_secondary_series)
                        else SECONDARY_FACE_COLOR
                        if (SECONDARY_IDX == 1)
                        else PRIMARY_FACE_COLOR
                        if SECONDARY_IDX == 2
                        else SECONDARY_FACE_COLOR
                    )
                    plt.scatter(
                        x,
                        y,
                        color=series_color,
                        facecolors=FACE_COLOR,
                        MARKER=MARKER,
                        label=legend,
                        s=plot_info.marker_size,
                    )

                    if has_y_error or has_x_error:
                        y_error = (
                            np.array(list(zip(y_neg_errors, y_pos_errors))).T
                            if has_y_error
                            else None
                        )
                        x_error = (
                            np.array(list(zip(x_neg_errors, x_pos_errors))).T
                            if has_x_error
                            else None
                        )
                        plt.errorbar(
                            x,
                            y,
                            yerr=y_error,
                            xerr=x_error,
                            fmt=ERROR_MARKER,
                            CAP_SIZE=CAP_SIZE,
                            ecolor=series_color,
                            label="_nolegend_",
                        )
                elif plot_info.plot_type == PlotType.LINE:
                    x, y = (
                        plot_data[series_name][constants.X_HEADER],
                        plot_data[series_name][constants.Y_HEADER],
                    )
                    if plot_info.sort_data:
                        x = np.array(x)
                        y = np.array(y)
                        sorting = np.argsort(x)
                        x = x[sorting]
                        y = y[sorting]

                    plt.plot(x, y, color=series_color, label=legend)

                    if has_y_error:
                        plt.fill_between(
                            x,
                            np.array(y) - np.array(y_neg_errors),
                            np.array(y) + np.array(y_pos_errors),
                            color=series_color,
                            alpha=0.25,
                            label="_nolegend_",
                        )
                    if has_x_error:
                        plt.fill_between(
                            y,
                            np.array(x) - np.array(x_neg_errors),
                            np.array(x) + np.array(x_pos_errors),
                            color=series_color,
                            alpha=0.25,
                            label="_nolegend_",
                        )

                if plot_info.draw_line_at_maximum:
                    # draw a vertical line that passes through the
                    # maximum y-value of the series
                    if len(x) > 0:
                        yMin, yMax = np.min(y), np.max(y)
                        xMax = x[np.nanargmax(y)]
                        if plot_info.draw_line_at_maximum:
                            plt.vlines(
                                xMax,
                                min(0, yMin),
                                max(0, yMax),
                                linestyles="dashed",
                                label="_nolegend_",
                            )

            for i, fit_line in enumerate(plot_info.fit_lines):
                # draw lines on the figure defined by y = fit_function(x)
                # with a domain of [fit_start_x, fit_end_x]
                fit_start_x, fit_end_x, fit_function = fit_line
                fit_x = np.linspace(fit_start_x, fit_end_x)
                fit_y = list(map(fit_function, fit_x))
                plt.plot(
                    fit_x,
                    fit_y,
                    color=FIT_LINE_COLOR
                    if not plot_info.use_color_map_for_fit_color
                    else COLOR_SCHEME[i],
                    linestyle=FIT_LINE_STYLE,
                )

            if plot_info.legends:
                # include series legends based on specified formatting
                if plot_info.reverse_legends:
                    a = plt.gca().get_legend_handles_labels()
                    handles, labels = plt.gca().get_legend_handles_labels()
                    plt.gca().legend(
                        handles[::-1],
                        labels[::-1],
                        fontsize=plot_info.legend_size,
                        loc=plot_info.legend_loc,
                        ncol=plot_info.legend_n_cols,
                        bbox_to_anchor=plot_info.legend_bbox_anchor,
                    )
                else:
                    plt.legend(
                        fontsize=plot_info.legend_size,
                        loc=plot_info.legend_loc,
                        ncol=plot_info.legend_n_cols,
                        bbox_to_anchor=plot_info.legend_bbox_anchor,
                    )

            if plot_info.vline_at_x is not None:
                # draw vertical line that passes through a specified x-value
                plt.vlines(
                    plot_info.vline_at_x,
                    0 if np.sign(MIN_Y) == np.sign(MAX_Y) else MIN_Y,
                    MAX_Y if MAX_Y > 0 else MIN_Y,
                    linestyles=DASHED_LINE_STYLE,
                    COLOR_SCHEME=DASHED_LINE_COLOR,
                    label="_nolegend_",
                )
            if plot_info.hline_at_y is not None:
                # draw horizontal line that passes through a specified y-value
                plt.hlines(
                    plot_info.hline_at_y,
                    0 if np.sign(MIN_X) == np.sign(MAX_X) else MIN_X,
                    MAX_X + (SERIES_COUNT * plot_info.secondary_series_stagger)
                    if MAX_X > 0
                    else MIN_X,
                    linestyles=DASHED_LINE_STYLE,
                    COLOR_SCHEME=DASHED_LINE_COLOR,
                    label="_nolegend_",
                )

            # convert axes to logarithm scale based on specified formatting
            if plot_info.yaxis_log_scale:
                plt.yscale("log", base=10)
            if plot_info.xaxis_log_scale:
                plt.xscale("log", base=10)

            # additional figure formatting
            plt.ylim(plot_info.ylim)
            plt.xlim(plot_info.xlim)
            plt.title(plot_info.title, loc="left")
            plt.xlabel(plot_info.xlabel, fontsize=plot_info.font_size)
            plt.ylabel(plot_info.ylabel, fontsize=plot_info.font_size)
            plt.xticks(fontsize=plot_info.font_size)
            plt.yticks(fontsize=plot_info.font_size)
            plt.gca().tick_params(axis="both", which="both", direction="in")
            if not plot_info.show_x_axis_numbers:
                plt.gca().axes.xaxis.set_ticklabels([])
            if not plot_info.show_y_axis_numbers:
                plt.gca().axes.yaxis.set_ticklabels([])
            if plot_info.figure_size_in_inches:
                plt.gcf().set_size_inches(*plot_info.figure_size_in_inches)

            # save figure using specified file extension
            if (
                plot_info.suplot_tuple is None
                or (plot_info.suplot_tuple[0] * plot_info.suplot_tuple[1])
                <= plot_info.suplot_tuple[2]
            ):
                saveFolder = create_folder_if_not_exists(
                    join(root_folder, plot_info.save_sub_directory)
                )
                saveDestination = join(saveFolder, sheetname + plot_info.save_file_extension)
                plt.savefig(saveDestination)
                plt.close()
